from openpyxl import load_workbook
import csv

def process_file(filepath, filename, output_path):
    workbook = load_workbook(filename=filepath + filename, read_only=True)
    sheets = workbook.sheetnames

    for sheet in sheets:
        print('Work with sheet [' + sheet + ']')

        # 0. Activate sheet
        worksheet = workbook[sheet]

        # 1. Create output list
        output_list = []

        # Latitude is first row
        # Longutide is first column

        # Create support list with longitude
        long_list = []

        for lat_row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if lat_row_idx == 0:
                # In this case we assume what we work with header which includes
                # longitudes
                for idx in range(1, len(row)):
                    long_list.append(row[idx])
                continue
        
            current_lat = None
            for long_column_idx, value in enumerate(row):
                if long_column_idx == 0:
                    # In this case we assume that in first column exist a latitude list
                    current_lat = value
                    continue

                # Except of first raw and first column - all remains cells contain data
                output_list.append([
                    current_lat,                        # latitude
                    long_list[long_column_idx - 1],     # longitude
                    value                               # value
                ])

        # 3. Write values to CSV
        with open(output_path + filename + '_' + sheet + '.txt', 'w', newline='') as f:
            writer = csv.writer(f, delimiter='\t')
            for line in output_list:
                writer.writerow(line)

def main():
    print("Scrript is started")

    # TODO change file here
    process_file("./input/", "V wind 975 hPa march 2020.xlsx", "./output/")
    
    print('Script is ended')

if __name__ == "__main__":
    main()